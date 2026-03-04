"""
ResultOps - Excel Export Service
Generates multi-sheet Excel workbooks for result reports.
"""

import io
import logging
<<<<<<< HEAD
from typing import Optional
=======

>>>>>>> origin/develop

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from analytics.analytics import Analytics

logger = logging.getLogger(__name__)

<<<<<<< HEAD
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL     = PatternFill("solid", fgColor="D6E4F7")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_BORDER_SIDE  = Side(style="thin", color="AAAAAA")
_CELL_BORDER  = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE, bottom=_BORDER_SIDE
=======
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL = PatternFill("solid", fgColor="D6E4F7")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BORDER_SIDE = Side(style="thin", color="AAAAAA")
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE
>>>>>>> origin/develop
)


class ExcelExportService:
    """Generates Excel reports for a semester."""

    def __init__(self):
        self.analytics = Analytics()

<<<<<<< HEAD
    def generate_excel(self, semester_id: str, title: str = "ResultOps Report") -> bytes:
=======
    def generate_excel(
        self, semester_id: str, title: str = "ResultOps Report"
    ) -> bytes:
>>>>>>> origin/develop
        """
        Generate a complete Excel workbook with:
        - Rank List sheet
        - Subject Analytics sheet
        - SGPA Distribution sheet
<<<<<<< HEAD
        
=======

>>>>>>> origin/develop
        Returns:
            Raw bytes of the .xlsx file.
        """
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Sheet 1: Rank List
            rank_df = self.analytics.student_rank_list(semester_id)
            if not rank_df.empty:
                rank_df.to_excel(writer, sheet_name="Rank List", index=False)
                _style_sheet(writer.sheets["Rank List"], rank_df)

            # Sheet 2: Subject Analytics
            subj_df = self.analytics.subject_analytics(semester_id)
            if not subj_df.empty:
                subj_df.to_excel(writer, sheet_name="Subject Analytics", index=False)
                _style_sheet(writer.sheets["Subject Analytics"], subj_df)

            # Sheet 3: SGPA Distribution
            dist_df = self.analytics.sgpa_distribution(semester_id)
            if not dist_df.empty:
                dist_df.to_excel(writer, sheet_name="SGPA Distribution", index=False)
                _style_sheet(writer.sheets["SGPA Distribution"], dist_df)

            # Sheet 4: Summary
            summary = self.analytics.semester_summary(semester_id)
            if summary:
                summary_df = pd.DataFrame(
                    list(summary.items()), columns=["Metric", "Value"]
                )
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                _style_sheet(writer.sheets["Summary"], summary_df)

        output.seek(0)
        return output.read()


def _style_sheet(ws, df: pd.DataFrame) -> None:
    """Apply formatting to a worksheet: header row + alternating row colors."""
    # Style header row
    for col_num, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_num)
<<<<<<< HEAD
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
=======
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
>>>>>>> origin/develop
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _CELL_BORDER

    # Style data rows
    for row_num in range(2, ws.max_row + 1):
        fill = _ALT_FILL if row_num % 2 == 0 else PatternFill()
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
<<<<<<< HEAD
            cell.fill   = fill
=======
            cell.fill = fill
>>>>>>> origin/develop
            cell.border = _CELL_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit column widths
    for col_num, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_num)
        max_len = max(
            len(str(col_name)),
<<<<<<< HEAD
            *[len(str(ws.cell(row=r, column=col_num).value or "")) for r in range(2, ws.max_row + 1)],
            0
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
=======
            *[
                len(str(ws.cell(row=r, column=col_num).value or ""))
                for r in range(2, ws.max_row + 1)
            ],
            0
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_excel(metadata, students) -> bytes:
    """
    Standalone helper used by upload_page.py.
    Builds an Excel report directly from in-memory parsed objects
    (parser.refactored_parser.PDFMetadata + list[StudentRecord])
    without querying Firebase.

    Returns raw bytes of the .xlsx file.
    """
    output = io.BytesIO()

    # ── Build student summary DataFrame ─────────────────────────────────────
    student_rows = []
    for rank, s in enumerate(
        sorted(students, key=lambda x: (x.sgpa or 0), reverse=True), start=1
    ):
        student_rows.append(
            {
                "Rank": rank,
                "PRN": s.prn,
                "Seat No": s.seat_no,
                "Name": s.name,
                "SGPA": s.sgpa,
                "Credits Earned": s.credits_earned,
                "Credits Total": s.credits_total,
                "Status": s.status,
                "Subjects": len(s.subjects),
            }
        )
    rank_df = pd.DataFrame(student_rows)

    # ── Build subject analytics DataFrame ───────────────────────────────────
    subject_map: dict = {}
    for s in students:
        for subj in s.subjects:
            code = subj.code
            subject_map.setdefault(code, []).append(subj)

    subj_rows = []
    for code, entries in sorted(subject_map.items()):
        totals = [e.total for e in entries if e.total is not None]
        grades = [e.grade for e in entries]
        passed = sum(1 for g in grades if g and g not in ("F", "FF", "AB"))
        appeared = len(entries)
        subj_rows.append(
            {
                "Subject Code": code,
                "Appeared": appeared,
                "Passed": passed,
                "Failed": appeared - passed,
                "Pass %": round(passed / appeared * 100, 1) if appeared else 0,
                "Highest": max(totals, default=0),
                "Average": round(sum(totals) / len(totals), 2) if totals else 0,
            }
        )
    subj_df = pd.DataFrame(subj_rows)

    # ── Build summary DataFrame ──────────────────────────────────────────────
    sgpas = [s.sgpa for s in students if s.sgpa is not None]
    statuses = [s.status for s in students]
    summary_data = {
        "University": getattr(metadata, "university", ""),
        "College": getattr(metadata, "college", ""),
        "Department": getattr(metadata, "department", ""),
        "Semester": getattr(metadata, "semester", ""),
        "Session": getattr(metadata, "session", ""),
        "Total Students": len(students),
        "Avg SGPA": round(sum(sgpas) / len(sgpas), 2) if sgpas else 0,
        "Max SGPA": max(sgpas, default=0),
        "Min SGPA": min(sgpas, default=0),
        "Pass Count": statuses.count("PASS"),
        "Fail Count": statuses.count("FAIL"),
        "Pass %": (
            round(statuses.count("PASS") / len(statuses) * 100, 1) if statuses else 0
        ),
    }
    summary_df = pd.DataFrame(list(summary_data.items()), columns=["Metric", "Value"])

    # ── Write sheets ─────────────────────────────────────────────────────────
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if not rank_df.empty:
            rank_df.to_excel(writer, sheet_name="Rank List", index=False)
            _style_sheet(writer.sheets["Rank List"], rank_df)

        if not subj_df.empty:
            subj_df.to_excel(writer, sheet_name="Subject Analytics", index=False)
            _style_sheet(writer.sheets["Subject Analytics"], subj_df)

        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        _style_sheet(writer.sheets["Summary"], summary_df)

    output.seek(0)
    return output.read()
>>>>>>> origin/develop
