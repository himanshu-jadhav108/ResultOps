"""
Upload & Parse page — uses the robust metadata/student parsers.
Includes parsing confidence score and enhanced progress UX.
"""

import time
import io
import logging

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.auth import auth_manager
from utils.theme import theme_manager

logger = logging.getLogger(__name__)


# ── Parsing Confidence Score ──────────────────────────────────────────────────


def _compute_confidence(metadata, students, validation) -> tuple[int, list[str]]:
    """
    Compute a 0-100 parsing confidence score.
    Returns (score, list_of_issues).
    """
    score = 100
    issues = []

    # Metadata completeness (up to -25)
    if not metadata.university_name or metadata.university_name in ("", "Unknown"):
        score -= 8
        issues.append("University name not detected")
    if not metadata.college_name or metadata.college_name in ("", "Unknown"):
        score -= 5
        issues.append("College name not detected")
    if not metadata.department_name or metadata.department_name in ("", "Unknown"):
        score -= 7
        issues.append("Department name not detected")
    if not metadata.semester_number:
        score -= 5
        issues.append("Semester number not detected")

    if not students:
        return 0, ["No student records parsed"]

    # Student data completeness (up to -35)
    total = len(students)
    missing_sgpa = sum(1 for s in students if s.sgpa is None)
    missing_name = sum(1 for s in students if not s.name)
    missing_prn = sum(1 for s in students if not s.prn)
    no_subjects = sum(1 for s in students if not s.subjects)

    if missing_sgpa > 0:
        pct = missing_sgpa / total
        deduct = int(pct * 20)
        score -= deduct
        issues.append(f"{missing_sgpa}/{total} students missing SGPA")
    if missing_name > 0:
        score -= min(int(missing_name / total * 10), 10)
        issues.append(f"{missing_name}/{total} students missing name")
    if missing_prn > 0:
        score -= min(int(missing_prn / total * 10), 10)
        issues.append(f"{missing_prn}/{total} students missing PRN")
    if no_subjects > 0:
        pct = no_subjects / total
        score -= int(pct * 15)
        issues.append(f"{no_subjects}/{total} students have no subject data")

    # Validation warnings (up to -20)
    warn_lines = [line for line in validation.summary_lines() if line.startswith("⚠️")]
    error_lines = [line for line in validation.summary_lines() if line.startswith("❌")]
    score -= min(len(warn_lines) * 3, 10)
    score -= min(len(error_lines) * 5, 10)
    if warn_lines:
        issues.append(f"{len(warn_lines)} validation warning(s)")
    if error_lines:
        issues.append(f"{len(error_lines)} validation error(s)")

    return max(0, min(100, score)), issues


def _render_confidence(score: int, issues: list[str]):
    """Render a colour-coded confidence score bar. Fully theme-aware."""
    c = theme_manager.colors
    if score >= 85:
        color, label = c["success"], "Excellent"
    elif score >= 65:
        color, label = c["warning"], "Good"
    elif score >= 40:
        color, label = "#f97316", "Fair"  # Orange
    else:
        color, label = c["error"], "Poor"

    issues_html = (
        "<ul style='margin-top:12px;padding-left:18px;line-height:1.6;font-size:13px;'>"
        + "".join(
            f"<li style='color:{c['text_sub']};margin-bottom:4px;'>{i}</li>"
            for i in issues
        )
        + "</ul>"
        if issues
        else f"<div style='color:{c['success']};font-size:13px;margin-top:10px;font-weight:600;'>✅ No issues detected — data looks complete.</div>"
    )

    st.markdown(
        f"""
<div class="premium-card" style="border-left: 5px solid {color};">
  <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:14px;">
    <span style="font-size:14px; font-weight:700; color:{c['text_muted']}; text-transform:uppercase; letter-spacing:1px;"
    >🎯 Parsing Confidence</span>
    <span style="font-size:20px; font-weight:800; color:{color};">{score}% — {label}</span>
  </div>
  <div style="background: {c['card_border']}; border-radius:100px; height:8px; overflow:hidden;">
    <div style="background:{color}; width:{score}%; height:100%; border-radius:100px; 
                transition:width 1s cubic-bezier(0.4, 0, 0.2, 1); box-shadow: 0 0 12px {color}44;"></div>
  </div>
  {issues_html}
</div>""",
        unsafe_allow_html=True,
    )


def _sgpa_category(sgpa):
    if sgpa is None:
        return "N/A"
    if sgpa >= 7.75:
        return "Distinction"
    if sgpa >= 6.75:
        return "First Class"
    if sgpa >= 6.0:
        return "Higher Second"
    if sgpa >= 5.0:
        return "Second Class"
    if sgpa >= 4.0:
        return "Pass"
    return "Fail"


def _excel_filename(metadata):
    dept = metadata.department_name.replace(" ", "_")[:25]
    return f"ResultOps_{dept}_Sem{metadata.semester_number}_{metadata.session_type}{metadata.session_year}.xlsx"


def _generate_excel(metadata, students) -> bytes:
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    ALT_FILL = PatternFill("solid", fgColor="EBF2FF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER_SIDE = Side(style="thin", color="C5D5EA")
    CELL_BORDER = Border(
        left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE
    )

    def style_ws(ws, df):
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = CELL_BORDER
        for row_num in range(2, ws.max_row + 1):
            fill = ALT_FILL if row_num % 2 == 0 else PatternFill()
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = fill
                cell.border = CELL_BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_num, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_num)
            max_len = max(
                len(str(col_name)),
                *[
                    len(str(ws.cell(row=r, column=col_num).value or ""))
                    for r in range(2, ws.max_row + 1)
                ],
                0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)
        ws.row_dimensions[1].height = 28

    output = io.BytesIO()
    subj_map = {}
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df1 = pd.DataFrame(
            [
                {
                    "PRN": s.prn,
                    "Seat No": s.seat_no,
                    "Name": s.name,
                    "SGPA": s.sgpa,
                    "Credits Earned": s.credits_earned,
                    "Credits Total": s.credits_total,
                    "Subjects": len(s.subjects),
                    "Status": "PASS" if (s.sgpa or 0) >= 4.0 else "FAIL",
                    "Category": _sgpa_category(s.sgpa),
                }
                for s in students
            ]
        )
        df1.to_excel(writer, sheet_name="Student Master", index=False)
        style_ws(writer.sheets["Student Master"], df1)

        sorted_s = sorted(students, key=lambda s: s.sgpa or 0, reverse=True)
        df2 = pd.DataFrame(
            [
                {
                    "Rank": i,
                    "PRN": s.prn,
                    "Seat No": s.seat_no,
                    "Name": s.name,
                    "SGPA": s.sgpa,
                    "Status": "PASS" if (s.sgpa or 0) >= 4.0 else "FAIL",
                    "Category": _sgpa_category(s.sgpa),
                }
                for i, s in enumerate(sorted_s, 1)
            ]
        )
        df2.to_excel(writer, sheet_name="Rank List", index=False)
        style_ws(writer.sheets["Rank List"], df2)

        for s in students:
            for subj in s.subjects:
                subj_map.setdefault(subj.subject_code, []).append(subj)

        subj_rows = []
        for code, entries in sorted(subj_map.items()):
            totals = [e.total for e in entries if e.total is not None]
            grades = [e.grade for e in entries]
            passed = sum(1 for g in grades if g not in ("F", "FF", "AB", None))
            appeared = len(entries)
            subj_rows.append(
                {
                    "Subject Code": code,
                    "Appeared": appeared,
                    "Passed": passed,
                    "Failed": appeared - passed,
                    "Pass %": round(passed / appeared * 100, 1) if appeared else 0,
                    "Highest": max(totals, default=0),
                    "Lowest": min(totals, default=0),
                    "Average": round(sum(totals) / len(totals), 2) if totals else 0,
                }
            )
        df3 = pd.DataFrame(subj_rows)
        if not df3.empty:
            df3.to_excel(writer, sheet_name="Subject Analytics", index=False)
            style_ws(writer.sheets["Subject Analytics"], df3)

        bins = [0, 4, 5, 6, 6.75, 7.75, 8.5, 10]
        labels = [
            "<4 (Fail)",
            "4-5 (Pass)",
            "5-6 (Second)",
            "6-6.75 (Higher Second)",
            "6.75-7.75 (First Class)",
            "7.75-8.5 (Distinction)",
            ">8.5 (Outstanding)",
        ]
        sgpas = [s.sgpa for s in students if s.sgpa is not None]
        dist = {lbl: 0 for lbl in labels}
        for sgpa in sgpas:
            for i in range(len(bins) - 1):
                if bins[i] <= sgpa < bins[i + 1]:
                    dist[labels[i]] += 1
                    break
            else:
                dist[labels[-1]] += 1
        df4 = pd.DataFrame(list(dist.items()), columns=["Category", "Count"])
        df4.to_excel(writer, sheet_name="SGPA Distribution", index=False)
        style_ws(writer.sheets["SGPA Distribution"], df4)

        sgpa_vals = [s.sgpa for s in students if s.sgpa is not None]
        pass_count = sum(1 for s in students if (s.sgpa or 0) >= 4.0)
        df5 = pd.DataFrame(
            [
                ("University", metadata.university_name),
                ("College", metadata.college_name),
                ("Department", metadata.department_name),
                ("Semester", metadata.semester_number),
                ("Session", f"{metadata.session_type} {metadata.session_year}"),
                ("", ""),
                ("Total Students", len(students)),
                (
                    "Average SGPA",
                    round(sum(sgpa_vals) / len(sgpa_vals), 2) if sgpa_vals else 0,
                ),
                ("Highest SGPA", max(sgpa_vals, default=0)),
                ("Lowest SGPA", min(sgpa_vals, default=0)),
                ("Pass Count", pass_count),
                ("Fail Count", len(students) - pass_count),
                (
                    "Pass %",
                    round(pass_count / len(students) * 100, 1) if students else 0,
                ),
                ("Distinctions (>=7.75)", sum(1 for v in sgpa_vals if v >= 7.75)),
                (
                    "First Class (6.75-7.75)",
                    sum(1 for v in sgpa_vals if 6.75 <= v < 7.75),
                ),
                ("Total Subjects", len(subj_map)),
            ],
            columns=["Metric", "Value"],
        )
        df5.to_excel(writer, sheet_name="Summary", index=False)
        style_ws(writer.sheets["Summary"], df5)

    output.seek(0)
    return output.read()


def render():
    c = theme_manager.colors
    st.title("📤 Upload University Result PDF")
    st.markdown(
        "Upload an **official university ledger PDF** (text‑based, not scanned). "
        "ResultOps auto‑detects all metadata and parses every student record."
    )
    st.markdown("---")

    uploaded_file = st.file_uploader(
        "📂 Choose PDF file",
        type=["pdf"],
        help="Must be a text-based university ledger PDF (not a scanned image).",
    )

    if not uploaded_file:
        st.markdown(
            f"""
        <div style="text-align:center; padding:48px; color:{c['text_muted']};">
            <div style="font-size:52px; margin-bottom:12px;">📄</div>
            <div style="font-size:18px; font-weight:600; color:{c['text_sub']};">No file uploaded yet</div>
            <div style="font-size:13px; margin-top:8px;">
                Drag &amp; drop or click Browse above to get started
            </div>
        </div>
        """,
            unsafe_allow_html=True,
        )
        return

    pdf_bytes = uploaded_file.read()

    # ── Parse ──────────────────────────────────────────────────────────────────
    progress = st.progress(0, text="⏳ Initialising parser…")
    try:
        from parser.pdf_parser import extract_text_from_pdf
        from parser.metadata_extractor import extract_metadata
        from parser.student_parser import parse_students
        from utils.validators import validate_students

        progress.progress(15, text="📖 Stage 1/4 — Extracting text from PDF…")
        full_text = extract_text_from_pdf(pdf_bytes)

        progress.progress(
            40, text="🔍 Stage 2/4 — Detecting metadata (University, Dept, Semester)…"
        )
        metadata = extract_metadata(full_text)

        progress.progress(70, text="👥 Stage 3/4 — Parsing student records…")
        students = parse_students(full_text)

        progress.progress(90, text="✅ Stage 4/4 — Validating data…")
        validation = validate_students(
            students, expected_semester=metadata.semester_number
        )

        progress.progress(100, text=f"✅ Done! Parsed {len(students)} students.")
        time.sleep(0.4)
        progress.empty()

    except Exception as e:
        progress.empty()
        st.error(f"❌ Parsing failed: {e}")
        return

    # ── Metadata cards ─────────────────────────────────────────────────────────
    st.markdown("### 🔍 Detected Metadata")
    c1, c2, c3, c4, c5 = st.columns(5)
    cards = [
        (c1, "#2d8cff", "🏛️ University", metadata.university_name),
        (c2, "#2d8cff", "🏫 College", metadata.college_name),
        (c3, "#f5a623", "📚 Department", metadata.department_name),
        (c4, "#00d97e", "📋 Semester", f"Semester {metadata.semester_number}"),
        (
            c5,
            "#a78bfa",
            "📅 Session",
            f"{metadata.session_type} {metadata.session_year}",
        ),
    ]
    for col, border_color, label, value in cards:
        with col:
            st.markdown(
                f"""
            <div class="premium-card" style="border-top:3px solid {border_color}; padding: 18px;">
                <div style="font-size:10px; text-transform:uppercase; color:{c['text_muted']}; 
                            letter-spacing:1px; font-weight:700; margin-bottom:10px;">{label}</div>
                <div style="font-size:14px; font-weight:700; color:{c['text']}; line-height:1.4;">{value}</div>
            </div>""",
                unsafe_allow_html=True,
            )

    st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("### 🧪 Validation Report")
    # validation already computed during parse stage above

    # ── Confidence Score ──────────────────────────────────────────────────────
    conf_score, conf_issues = _compute_confidence(metadata, students, validation)
    _render_confidence(conf_score, conf_issues)

    v1, v2 = st.columns([3, 2])
    with v1:
        warn_count = sum(
            1 for line in validation.summary_lines() if line.startswith("⚠️")
        )
        err_count = sum(
            1 for line in validation.summary_lines() if line.startswith("❌")
        )
        if warn_count or err_count:
            st.caption(
                f"⚠️ {warn_count} warning(s) · ❌ {err_count} error(s) detected during validation"
            )
        for line in validation.summary_lines():
            if line.startswith("✅"):
                st.success(line)
            elif line.startswith("⚠️"):
                st.warning(line)
            elif line.startswith("❌"):
                st.error(line)

    with v2:
        if validation.is_valid:
            st.markdown(
                f"""
            <div style="background:{c['card']}; border:2px solid {c['success']};
                        border-radius:12px; padding:28px; text-align:center;">
                <div style="font-size:40px;">✅</div>
                <div style="font-size:17px; font-weight:700; color:{c['success']}; margin-top:10px;">
                    Validation Passed</div>
                <div style="color:{c['success']}; opacity:0.8; margin-top:6px; font-size:13px;">
                    Ready to save to database</div>
            </div>""",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f"""
            <div style="background:{c['card']}; border:2px solid {c['error']};
                        border-radius:12px; padding:28px; text-align:center;">
                <div style="font-size:40px;">❌</div>
                <div style="font-size:17px; font-weight:700; color:{c['error']}; margin-top:10px;">
                    Validation Failed</div>
                <div style="color:{c['error']}; opacity:0.8; margin-top:6px; font-size:13px;">
                    Fix errors before saving</div>
            </div>""",
                unsafe_allow_html=True,
            )

    st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

    # ── Student Preview ────────────────────────────────────────────────────────
    if students:
        st.markdown("---")
        st.markdown(f"### 👥 Student Preview — {len(students)} records detected")
        preview_rows = [
            {
                "PRN": s.prn,
                "Seat No": s.seat_no,
                "Name": s.name,
                "SGPA": s.sgpa,
                "Subjects": len(s.subjects),
                "Status": "✅ PASS" if (s.sgpa or 0) >= 4.0 else "❌ FAIL",
            }
            for s in students[:15]
        ]
        st.dataframe(
            pd.DataFrame(preview_rows), use_container_width=True, hide_index=True
        )
        if len(students) > 15:
            st.caption(f"Showing first 15 of {len(students)} students.")

    # ── Download (no auth required) ────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Download Analysed Report")
    st.caption("Get the full Excel report now — no database save required.")

    dl_left, dl_right = st.columns([2, 3])
    with dl_left:
        try:
            excel_data = _generate_excel(metadata, students)
            st.download_button(
                label="⬇️ Download Excel Report (5 Sheets)",
                data=excel_data,
                file_name=_excel_filename(metadata),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.warning(f"Could not generate report: {e}")
    with dl_right:
        st.markdown(
            f"""
        <div style="background:{c['card']}; border:1px solid {c['card_border']};
                    border-radius:10px; padding:14px 18px; font-size:13px; color:{c['text_sub']};">
            <b style="color:{c['text']};">Report includes 5 sheets:</b><br>
            📋 Student Master &nbsp;·&nbsp; 🏆 Rank List &nbsp;·&nbsp;
            📚 Subject Analytics<br>📊 SGPA Distribution &nbsp;·&nbsp; 📝 Summary
        </div>""",
            unsafe_allow_html=True,
        )

    # ── Save to DB (WRITE auth required) ──────────────────────────────────────
    st.markdown("---")

    if not validation.is_valid:
        st.error("⛔ Cannot save — fix the validation errors above first.")
        return

    if not auth_manager.require_write_auth(show_ui=True):
        return

    save_col, note_col = st.columns([2, 3])
    with save_col:
        save_clicked = st.button(
            "💾 Save to ResultOps Database",
            type="primary",
            use_container_width=True,
        )
    with note_col:
        st.caption(
            "Saves all student records to Firebase. Duplicate semester uploads are blocked."
        )

    if save_clicked:
        from services.result_service import ResultService, DuplicateSemesterError

        try:
            with st.spinner("💾 Saving to Firebase — please wait..."):
                service = ResultService()
                summary = service.save_results(metadata, students)

            st.balloons()
            st.success("✅ Results saved successfully!")

            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Students", summary["students_inserted"])
            r2.metric("Results", summary["results_inserted"])
            r3.metric("Marks", summary["marks_inserted"])
            r4.metric("Semester", f"Sem {summary['semester']}")

            st.download_button(
                label="⬇️ Download Full Excel Report",
                data=_generate_excel(metadata, students),
                file_name=_excel_filename(metadata),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        except DuplicateSemesterError as e:
            st.error(f"🚫 Duplicate Upload Blocked — {e}")
            st.download_button(
                label="⬇️ Download Excel Report Anyway",
                data=_generate_excel(metadata, students),
                file_name=_excel_filename(metadata),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"❌ Database error: {e}")
            logger.exception("DB save error")
