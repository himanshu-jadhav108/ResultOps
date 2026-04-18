"""
Analytics Dashboard — enhanced with subject difficulty ranking, semester comparison,
proper tie-aware ranks, failed student CSV download, and Pass/Fail filter.
"""

import io
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.theme import theme_manager


def _try_load():
    try:
        from analytics.analytics import Analytics

        return Analytics()
    except Exception as e:
        st.error(f"❌ Cannot connect to Firebase: {e}")
        return None


def _plotly_cfg():
    return {"displayModeBar": False}


# ── Excel export — selective sheets ───────────────────────────────────────────
def _build_excel(
    summary, rank_df, subj_df, dist_df, include, meta_info=None, master_df=None
) -> bytes:
    """Build Excel with only the selected sheets."""
    HFILL = PatternFill("solid", fgColor="1F4E79")
    AFILL = PatternFill("solid", fgColor="EBF2FF")
    HFONT = Font(bold=True, color="FFFFFF", size=11)
    BSIDE = Side(style="thin", color="C5D5EA")
    BORD = Border(left=BSIDE, right=BSIDE, top=BSIDE, bottom=BSIDE)

    def _style(ws, df):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(1, col)
            cell.fill = HFILL
            cell.font = HFONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = BORD
        for r in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(r, col)
                cell.fill = AFILL if r % 2 == 0 else PatternFill()
                cell.border = BORD
                cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_idx, col_name in enumerate(df.columns, 1):
            ltr = get_column_letter(col_idx)
            mx = max(
                len(str(col_name)),
                *[
                    len(str(ws.cell(r, col_idx).value or ""))
                    for r in range(2, ws.max_row + 1)
                ],
                0,
            )
            ws.column_dimensions[ltr].width = min(mx + 4, 45)
        ws.row_dimensions[1].height = 28

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        if (
            include.get("student_master")
            and master_df is not None
            and not master_df.empty
        ):
            master_df.to_excel(w, sheet_name="Student Master", index=False)
            _style(w.sheets["Student Master"], master_df)

        if include.get("summary"):
            summary_data = {}
            if meta_info:
                summary_data["University"] = meta_info.get("university", "")
                summary_data["College"] = meta_info.get("college", "")
                summary_data["Department"] = meta_info.get("department", "")
                summary_data["Semester"] = meta_info.get("semester", "")
                summary_data["---"] = "---"
            summary_data.update(summary)
            s_df = pd.DataFrame(list(summary_data.items()), columns=["Metric", "Value"])
            s_df.to_excel(w, sheet_name="Summary", index=False)
            _style(w.sheets["Summary"], s_df)

        if include.get("rank_list") and not rank_df.empty:
            rank_df.to_excel(w, sheet_name="Rank List", index=False)
            _style(w.sheets["Rank List"], rank_df)

        if include.get("subject") and not subj_df.empty:
            subj_df.to_excel(w, sheet_name="Subject Analytics", index=False)
            _style(w.sheets["Subject Analytics"], subj_df)

        if include.get("sgpa_dist") and not dist_df.empty:
            dist_df.to_excel(w, sheet_name="SGPA Distribution", index=False)
            _style(w.sheets["SGPA Distribution"], dist_df)

    out.seek(0)
    return out.read()


# ── Theme-aware row highlight helpers ─────────────────────────────────────────
def _get_highlight_styles():
    if theme_manager.is_dark:
        return {
            "distinction": "background-color: rgba(16, 185, 129, 0.15); color: #10b981; font-weight:700",
            "first_class": "background-color: rgba(99, 102, 241, 0.15); color: #818cf8; font-weight:700",
            "fail": "background-color: rgba(239, 68, 68, 0.15); color: #f87171; font-weight:700",
            "pass_high": "color: #10b981; font-weight:800",
            "pass_mid": "color: #f59e0b; font-weight:800",
            "pass_low": "color: #ef4444; font-weight:800",
        }
    else:
        return {
            "distinction": "background-color: #ecfdf5; color: #059669; font-weight:700",
            "first_class": "background-color: #eef2ff; color: #4f46e5; font-weight:700",
            "fail": "background-color: #fef2f2; color: #dc2626; font-weight:700",
            "pass_high": "color: #059669; font-weight:800",
            "pass_mid": "color: #d97706; font-weight:800",
            "pass_low": "color: #dc2626; font-weight:800",
        }


# ── Plotly charts ──────────────────────────────────────────────────────────────


def _plotly_bar(df, x, y, title, color, height=300):
    tc = theme_manager.colors
    fig = px.bar(df, x=x, y=y, title=title, color_discrete_sequence=[color], text=y)
    fig.update_traces(textposition="outside", textfont_color=tc["text"])
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color=tc["text"], family="Plus Jakarta Sans, sans-serif"),
        title_font=dict(size=15, color=tc["text"], weight="bold"),
        margin=dict(t=50, b=80, l=10, r=10),
        height=height,
        xaxis=dict(
            showgrid=False,
            type="category",
            tickangle=-35,
            automargin=True,
            tickfont=dict(color=tc["plotly_tick"], size=11),
            title_font=dict(color=tc["text"]),
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor=tc["plotly_grid"],
            tickfont=dict(color=tc["plotly_tick"], size=11),
            title_font=dict(color=tc["text"]),
        ),
    )
    return fig


def _plotly_line(df, x, y_cols, title):
    tc = theme_manager.colors
    colors = ["#2d8cff", "#00d97e", "#f5a623", "#a78bfa"]
    fig = go.Figure()
    for i, col in enumerate(y_cols):
        fig.add_trace(
            go.Scatter(
                x=df[x],
                y=df[col],
                mode="lines+markers+text",
                name=col,
                text=[str(v) for v in df[col]],
                textposition="top center",
                textfont=dict(color=tc["text"]),
                line=dict(color=colors[i % len(colors)], width=2.5),
                marker=dict(size=8),
            )
        )
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color=tc["text"], family="Plus Jakarta Sans, sans-serif"),
        title_text=title,
        title_font=dict(size=15, color=tc["text"], weight="bold"),
        margin=dict(t=60, b=40, l=10, r=10),
        xaxis=dict(
            showgrid=False,
            tickfont=dict(color=tc["plotly_tick"], size=11),
            title_font=dict(color=tc["text"]),
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor=tc["plotly_grid"],
            tickfont=dict(color=tc["plotly_tick"], size=11),
            title_font=dict(color=tc["text"]),
        ),
        legend=dict(font=dict(color=tc["text"])),
    )
    return fig


# ══════════════════════════════════════════════════════════════════════════════
# MAIN RENDER
# ══════════════════════════════════════════════════════════════════════════════


def render():
    c = theme_manager.colors
    hl = _get_highlight_styles()
    st.title("📊 Analytics Dashboard")
    st.markdown("---")

    analytics = _try_load()
    if analytics is None:
        return

    # ── Filters ────────────────────────────────────────────────────────────────
    st.markdown("### 🔎 Select Data")
    f1, f2, f3, f4 = st.columns(4)

    universities = analytics.get_universities()
    uni_list = [u["name"] for u in universities]
    uni_select = f1.selectbox("🏛️ University", ["— All —"] + uni_list, key="dash_uni")
    uni_filter = uni_select if uni_select != "— All —" else None

    colleges = analytics.get_colleges(uni_filter)
    col_list = [c_["name"] for c_ in colleges]
    col_select = f2.selectbox("🏫 College", ["— All —"] + col_list, key="dash_col")
    col_filter = col_select if col_select != "— All —" else None

    departments = analytics.get_departments(col_filter)
    dept_list = [d["name"] for d in departments]

    if not dept_list:
        st.info("📂 No departments found. Upload a result PDF first.")
        return

    dept_select = f3.selectbox("📚 Department", dept_list, key="dash_dept")

    semester_key = None
    sem_label = ""
    semesters = analytics.get_semesters_for_department(dept_select)
    if semesters:
        sem_labels = {
            f"Sem {s['semester_number']} — {s.get('session_type', '')} {s.get('session_year', '')}": s[
                "id"
            ]
            for s in semesters
        }
        sem_select = f4.selectbox(
            "📋 Semester", list(sem_labels.keys()), key="dash_sem"
        )
        semester_key = sem_labels.get(sem_select)
        sem_label = sem_select

    if not semester_key:
        st.info("📋 Select a semester to load analytics.")
        return

    st.markdown("---")

    summary = analytics.semester_summary(semester_key)
    if not summary:
        st.warning("No data found for this semester.")
        return

    # ── Semester overview metrics ──────────────────────────────────────────────
    st.markdown("### 📈 Semester Overview")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("👥 Students", summary["total_students"])
    m2.metric("📊 Avg SGPA", summary["avg_sgpa"])
    m3.metric("⭐ Highest", summary["max_sgpa"])
    m4.metric("🏆 Distinctions", summary["distinctions"])
    m5.metric("✅ Passed", summary["pass_count"])
    m6.metric("📈 Pass %", f"{summary['pass_percentage']}%")

    st.markdown("---")

    # ── Fetch data ─────────────────────────────────────────────────────────────
    master_df = analytics.student_master_list(semester_key)
    rank_df = analytics.student_rank_list(semester_key)
    subj_df = analytics.subject_analytics(semester_key)
    dist_df = analytics.sgpa_distribution(semester_key)

    # Apply proper tie-aware ranking
    if not rank_df.empty:
        # Proper ranking is now handled inside analytics.student_rank_list()
        pass

    # Detect the actual subject code column name
    subj_code_col = None
    if not subj_df.empty:
        for possible in ["Subject Code", "subject_code", "Subject_Code"]:
            if possible in subj_df.columns:
                subj_code_col = possible
                break
        if subj_code_col is None:
            subj_code_col = subj_df.columns[0]

    # ── Download Excel ─────────────────────────────────────────────────────────
    with st.expander("⬇️ Download Excel Report — click to configure"):
        chosen_uni = uni_select if uni_select != "— All —" else "(All)"
        chosen_col = col_select if col_select != "— All —" else "(All)"
        st.markdown(
            f"**Report for:** 🏛️ {chosen_uni} → 🏫 {chosen_col} → "
            f"📚 {dept_select} → 📋 {sem_label}"
        )

        st.caption("Select which sections to include:")
        dc1, dc2, dc3, dc4, dc5 = st.columns(5)
        inc_master = dc1.checkbox("📄 Student Master", value=True, key="dl_master")
        inc_summary = dc2.checkbox("📋 Summary", value=True, key="dl_summary")
        inc_rank = dc3.checkbox("🏆 Rank List", value=True, key="dl_rank")
        inc_subject = dc4.checkbox("📚 Subject Analytics", value=True, key="dl_subject")
        inc_sgpa = dc5.checkbox("📉 SGPA Distribution", value=True, key="dl_sgpa")

        include = {
            "student_master": inc_master,
            "summary": inc_summary,
            "rank_list": inc_rank,
            "subject": inc_subject,
            "sgpa_dist": inc_sgpa,
        }

        meta_info = {
            "university": uni_select if uni_select != "— All —" else "All",
            "college": col_select if col_select != "— All —" else "All",
            "department": dept_select,
            "semester": sem_label,
        }

        if not any(include.values()):
            st.warning("Select at least one section to download.")
        else:
            try:
                excel_bytes = _build_excel(
                    summary, rank_df, subj_df, dist_df, include, meta_info, master_df
                )
                safe_dept = dept_select.replace(" ", "_")[:20]
                safe_sem = sem_label.replace(" ", "_").replace("—", "").strip()[:15]
                st.download_button(
                    label="⬇️ Download Selected Report",
                    data=excel_bytes,
                    file_name=f"ResultOps_{safe_dept}_{safe_sem}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
            except Exception as e:
                st.warning(f"Excel generation failed: {e}")

    st.markdown("---")

    # ── View mode ─────────────────────────────────────────────────────────────
    st.markdown("##### 👁️ View Mode")
    vb1, vb2, vb3, _vb = st.columns([1, 1, 1, 3])

    if "dash_view_mode" not in st.session_state:
        st.session_state.dash_view_mode = "both"

    with vb1:
        if st.button(
            "📊 Charts Only",
            use_container_width=True,
            type=(
                "primary"
                if st.session_state.dash_view_mode == "charts"
                else "secondary"
            ),
        ):
            st.session_state.dash_view_mode = "charts"
            st.rerun()
    with vb2:
        if st.button(
            "📋 Tables Only",
            use_container_width=True,
            type=(
                "primary"
                if st.session_state.dash_view_mode == "tables"
                else "secondary"
            ),
        ):
            st.session_state.dash_view_mode = "tables"
            st.rerun()
    with vb3:
        if st.button(
            "📊📋 Both",
            use_container_width=True,
            type=(
                "primary" if st.session_state.dash_view_mode == "both" else "secondary"
            ),
        ):
            st.session_state.dash_view_mode = "both"
            st.rerun()

    show_charts = st.session_state.dash_view_mode in ("charts", "both")
    show_tables = st.session_state.dash_view_mode in ("tables", "both")

    st.markdown("---")

    # ── SGPA Distribution ─────────────────────────────────────────────────────
    st.markdown("### 📉 SGPA Distribution")

    if not dist_df.empty:
        if show_charts:
            fig_dist = _plotly_bar(
                dist_df,
                "SGPA Range",
                "Count",
                "SGPA Distribution",
                c["accent"],
                height=280,
            )
            st.plotly_chart(fig_dist, use_container_width=True, config=_plotly_cfg())
        if show_tables:
            total = dist_df["Count"].sum()
            display_dist = dist_df.copy()
            display_dist["Percentage"] = display_dist["Count"].apply(
                lambda x: f"{round(x / total * 100, 1)}%" if total > 0 else "0%"
            )
            st.dataframe(display_dist, use_container_width=True, hide_index=True)
    else:
        st.info("No distribution data.")

    st.markdown("---")

    # ── Student Rankings ──────────────────────────────────────────────────────
    st.markdown("### 🏆 Student Rankings")

    # Pass/Fail filter
    pf_filter = st.radio(
        "Filter by status",
        ["All", "Pass Only", "Fail Only"],
        horizontal=True,
        key="rank_pf_filter",
        label_visibility="collapsed",
    )

    if not rank_df.empty:
        filtered_rank = rank_df.copy()
        if pf_filter == "Pass Only":
            filtered_rank = filtered_rank[filtered_rank["Status"] == "PASS"]
        elif pf_filter == "Fail Only":
            filtered_rank = filtered_rank[filtered_rank["Status"] == "FAIL"]

        top10 = filtered_rank.head(10)[
            ["Rank", "Name", "SGPA", "Status", "Category"]
        ].copy()

        def _highlight(row):
            cat = row.get("Category", "")
            if cat == "Distinction":
                return [hl["distinction"]] * len(row)
            if cat == "First Class":
                return [hl["first_class"]] * len(row)
            return [""] * len(row)

        col_left, col_right = st.columns([3, 2])

        with col_left:
            st.markdown("#### 🥇 Top 10 Students")
            st.dataframe(
                top10.style.apply(_highlight, axis=1),
                use_container_width=True,
                hide_index=True,
            )

        with col_right:
            if show_charts:
                st.markdown("#### 📊 Top 10 SGPA")
                fig_top = _plotly_bar(
                    top10.rename(columns={"Name": "Student"}),
                    "Student",
                    "SGPA",
                    "Top 10 Students",
                    c["accent"],
                    height=280,
                )
                st.plotly_chart(fig_top, use_container_width=True, config=_plotly_cfg())

        with st.expander("📋 Full Rank List — click to expand"):

            def _hl_full(row):
                if row["Status"] == "FAIL":
                    return [hl["fail"]] * len(row)
                if row["Category"] == "Distinction":
                    return [hl["distinction"]] * len(row)
                if row["Category"] == "First Class":
                    return [hl["first_class"]] * len(row)
                return [""] * len(row)

            st.dataframe(
                filtered_rank.style.apply(_hl_full, axis=1),
                use_container_width=True,
                hide_index=True,
            )

        fail_df = rank_df[rank_df["Status"] == "FAIL"]
        if not fail_df.empty:
            with st.expander(f"⚠️ Fail List — {len(fail_df)} students"):
                fail_display = fail_df[["Rank", "PRN", "Seat No", "Name", "SGPA"]]
                st.dataframe(
                    fail_display,
                    use_container_width=True,
                    hide_index=True,
                )
                # CSV download for failed students
                csv_bytes = fail_display.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇️ Download Fail List (CSV)",
                    data=csv_bytes,
                    file_name=f"FailList_{dept_select.replace(' ', '_')}_{sem_label.replace(' ', '_')}.csv",
                    mime="text/csv",
                    type="primary",
                )
        else:
            st.success("✅ No failed students this semester!")
    else:
        st.info("No rank data available.")

    st.markdown("---")

    # ── Subject Analytics ─────────────────────────────────────────────────────
    st.markdown("### 📚 Subject-wise Analytics")

    if not subj_df.empty and subj_code_col:
        if show_tables:

            def _color_pass(val):
                try:
                    v = float(val)
                    if v >= 80:
                        return hl["pass_high"]
                    if v >= 60:
                        return hl["pass_mid"]
                    return hl["pass_low"]
                except Exception:
                    return ""

            st.dataframe(
                subj_df.style.map(_color_pass, subset=["Pass %"]),
                use_container_width=True,
                hide_index=True,
            )

        if show_charts:
            ch1, ch2 = st.columns(2)
            with ch1:
                fig_pass = _plotly_bar(
                    subj_df,
                    subj_code_col,
                    "Pass %",
                    "Pass % by Subject",
                    c["accent"],
                )
                st.plotly_chart(
                    fig_pass, use_container_width=True, config=_plotly_cfg()
                )
            with ch2:
                fig_avg = _plotly_bar(
                    subj_df,
                    subj_code_col,
                    "Average",
                    "Average Marks by Subject",
                    "#f5a623",
                )
                st.plotly_chart(fig_avg, use_container_width=True, config=_plotly_cfg())
    else:
        st.info("No subject data available.")

    st.markdown("---")

    # ── Subject Difficulty Ranking ────────────────────────────────────────────
    st.markdown("### 🎯 Subject Difficulty Ranking")
    st.caption(
        "Difficulty score = (Fail% × 0.6) + ((100 − AvgMarks) × 0.4). "
        "Higher score = harder subject."
    )

    with st.spinner("Computing difficulty…"):
        diff_df = analytics.subject_difficulty(semester_key)

    if not diff_df.empty:
        if show_tables:

            def _color_diff(val):
                try:
                    v = float(val)
                    if v >= 60:
                        return "color:#ff4d6d;font-weight:700"
                    if v >= 35:
                        return "color:#f5a623;font-weight:600"
                    return "color:#00d97e;font-weight:600"
                except Exception:
                    return ""

            st.dataframe(
                diff_df.style.map(_color_diff, subset=["Difficulty"]),
                use_container_width=True,
                hide_index=True,
            )

        if show_charts:
            fig_diff = _plotly_bar(
                diff_df.head(10),
                "Subject Code",
                "Difficulty",
                "Top 10 Most Difficult Subjects",
                "#ff4d6d",
            )
            st.plotly_chart(fig_diff, use_container_width=True, config=_plotly_cfg())
    else:
        st.info("No difficulty data available.")

    st.markdown("---")

    # ── Semester Performance Comparison ───────────────────────────────────────
    st.markdown("### 📈 Semester Performance Comparison")
    st.caption("Compare multiple semesters for the same department.")

    all_sems = analytics.get_semesters_for_department(dept_select)
    if len(all_sems) < 2:
        st.info("Upload at least 2 semesters for this department to enable comparison.")
    else:
        sem_label_map = {
            f"Sem {s['semester_number']} — {s.get('session_type', '')} {s.get('session_year', '')}": s[
                "id"
            ]
            for s in all_sems
        }
        selected_sem_labels = st.multiselect(
            "Select semesters to compare",
            list(sem_label_map.keys()),
            default=list(sem_label_map.keys()),
            key="comp_sems",
        )
        selected_keys = [
            sem_label_map[label]
            for label in selected_sem_labels
            if label in sem_label_map
        ]

        if len(selected_keys) >= 2:
            with st.spinner("Loading comparison data…"):
                comp_df = analytics.compare_semesters(selected_keys)
                trend = analytics.get_trend_analysis(comp_df)

            if not comp_df.empty:
                # Trend KPIs
                tk1, tk2, tk3 = st.columns(3)
                tk1.metric("📈 SGPA Trend", trend.get("sgpa_trend", "—"))
                tk2.metric("🏆 Best Semester", trend.get("best_semester", "—"))
                tk3.metric("📉 Worst Semester", trend.get("worst_semester", "—"))

                st.markdown(
                    "<div style='margin-top:12px'></div>", unsafe_allow_html=True
                )

                if show_charts:
                    fig_comp = _plotly_line(
                        comp_df,
                        "Semester",
                        ["Avg SGPA", "Pass %"],
                        "📈 SGPA & Pass % Trend Across Semesters",
                    )
                    st.plotly_chart(
                        fig_comp, use_container_width=True, config=_plotly_cfg()
                    )

                if show_tables:
                    st.dataframe(comp_df, use_container_width=True, hide_index=True)
        elif len(selected_keys) == 1:
            st.info("Select at least 2 semesters to see a comparison.")
