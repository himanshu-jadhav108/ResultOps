"""
<<<<<<< HEAD
ResultOps - Main Streamlit Application (Firebase Edition)
=======
ResultOps - Main Application Entry Point (Firebase Edition)
>>>>>>> origin/develop
University-grade result processing platform.
"""

import logging
<<<<<<< HEAD
import os
import io
import time
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(name)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger("resultops")

# ── Page config ───────────────────────────────────────────────────────────────
=======
from pathlib import Path

import streamlit as st
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s | %(name)s | %(levelname)s | %(message)s"
)
logger = logging.getLogger("resultops")

# ── Must be FIRST Streamlit call ──────────────────────────────────────────────
>>>>>>> origin/develop
st.set_page_config(
    page_title="ResultOps",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

<<<<<<< HEAD
# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0a1628 0%, #0f2c59 60%, #1a4a8a 100%) !important;
}
[data-testid="stSidebar"] * { color: #e8f0fe !important; }
[data-testid="stSidebar"] hr { border-color: rgba(45,140,255,0.3) !important; }

/* Main block */
.main .block-container {
    padding-top: 1.8rem;
    padding-bottom: 3rem;
    max-width: 1200px;
}

/* Headings */
h1 { color: #e8f0fe !important; font-weight: 700 !important; font-size: 28px !important; margin-bottom: 6px !important; }
h2, h3 { color: #93b5e1 !important; font-weight: 600 !important; }

/* Metric cards */
[data-testid="stMetric"] {
    background: #0d1f3c;
    border: 1px solid rgba(45,140,255,0.25);
    border-radius: 12px;
    padding: 16px 18px;
}
[data-testid="stMetricLabel"] { color: #7a9cc4 !important; font-size: 12px !important; }
[data-testid="stMetricValue"] { color: #ffffff !important; font-weight: 700 !important; font-size: 22px !important; }

/* Primary button */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1a4a8a, #2d8cff) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    font-size: 15px !important;
    padding: 10px 28px !important;
    transition: opacity 0.2s !important;
}
.stButton > button[kind="primary"]:hover { opacity: 0.85 !important; }

/* Secondary button */
.stButton > button[kind="secondary"] {
    background: transparent !important;
    color: #2d8cff !important;
    border: 1px solid #2d8cff !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
}

/* Alerts */
.stAlert { border-radius: 10px !important; }

/* Dataframe */
[data-testid="stDataFrame"] { border-radius: 10px; overflow: hidden; }

/* Expander */
.streamlit-expanderHeader {
    background: #0d1f3c !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    color: #93b5e1 !important;
    border: 1px solid rgba(45,140,255,0.2) !important;
}

/* Divider */
hr { border: none !important; border-top: 1px solid rgba(45,140,255,0.15) !important; margin: 24px 0 !important; }

/* File uploader */
[data-testid="stFileUploader"] {
    border: 2px dashed rgba(45,140,255,0.4) !important;
    border-radius: 12px !important;
    background: #0d1f3c !important;
}

/* Selectbox */
[data-testid="stSelectbox"] > div > div {
    background: #0d1f3c !important;
    border: 1px solid rgba(45,140,255,0.3) !important;
    border-radius: 8px !important;
    color: #e8f0fe !important;
}

/* Progress bar */
[data-testid="stProgressBar"] > div { background: #2d8cff !important; }

/* Sidebar logo fallback */
.sidebar-logo {
    text-align: center;
    padding: 24px 10px 8px 10px;
    font-size: 26px;
    font-weight: 800;
    letter-spacing: 1px;
    color: white;
}
.sidebar-tagline {
    text-align: center;
    font-size: 11px;
    color: #7fa8d4;
    margin-bottom: 8px;
    letter-spacing: 0.5px;
}

/* Metadata card */
.meta-card {
    background: #0d1f3c;
    border: 1px solid rgba(45,140,255,0.25);
    border-radius: 12px;
    padding: 16px 18px;
    height: 100%;
}
.meta-label {
    font-size: 11px;
    color: #7a9cc4;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-bottom: 8px;
    font-weight: 500;
}
.meta-value {
    font-size: 14px;
    font-weight: 700;
    color: #ffffff;
    line-height: 1.5;
    word-break: break-word;
}
</style>
""", unsafe_allow_html=True)


# ==============================================================================
# SERVICES  ← MUST be at module level, NOT inside the sidebar block
# ==============================================================================

@st.cache_resource(show_spinner=False)
def load_services():
    from analytics.analytics import Analytics
    from services.excel_export import ExcelExportService
    return Analytics(), ExcelExportService()


def try_load_services():
    """Load services with a friendly error card on Firebase failure."""
    try:
        return load_services()
    except Exception as e:
        st.markdown(f"""
        <div style="background:#1a0a00; border:1px solid #f5a623;
                    border-left:4px solid #f5a623; border-radius:10px;
                    padding:20px; margin-top:10px;">
            <div style="color:#f5a623; font-weight:700; font-size:15px; margin-bottom:10px;">
                ⚠️ Cannot Connect to Firebase
            </div>
            <div style="color:#d4a96a; font-size:13px; line-height:1.8;">
                <b>Most likely causes:</b><br>
                &nbsp;&nbsp;• <code>firebase_key.json</code> not found in project folder<br>
                &nbsp;&nbsp;• Firestore not enabled in Firebase Console<br>
                &nbsp;&nbsp;• Internet / network issue<br>
                &nbsp;&nbsp;• Missing <code>[firebase]</code> section in Streamlit Cloud secrets<br><br>
                <b>Quick fix:</b> Run <code>python test_connection.py</code> to diagnose.
            </div>
            <details style="margin-top:10px;">
                <summary style="color:#7a9cc4; cursor:pointer; font-size:12px;">
                    Show error details
                </summary>
                <code style="font-size:11px; color:#ff4d6d;">{str(e)}</code>
            </details>
        </div>
        """, unsafe_allow_html=True)
        return None, None


# ==============================================================================
# HELPERS
# ==============================================================================

def _sgpa_category(sgpa):
    if sgpa is None:  return "N/A"
    if sgpa >= 7.75:  return "Distinction"
    if sgpa >= 6.75:  return "First Class"
    if sgpa >= 6.0:   return "Higher Second"
    if sgpa >= 5.0:   return "Second Class"
    if sgpa >= 4.0:   return "Pass"
    return "Fail"


def _excel_filename(metadata):
    dept = metadata.department_name.replace(" ", "_")[:25]
    return f"ResultOps_{dept}_Sem{metadata.semester_number}_{metadata.session_type}{metadata.session_year}.xlsx"


def _generate_excel_report(metadata, students) -> bytes:
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    ALT_FILL    = PatternFill("solid", fgColor="EBF2FF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER_SIDE = Side(style="thin", color="C5D5EA")
    CELL_BORDER = Border(
        left=BORDER_SIDE, right=BORDER_SIDE,
        top=BORDER_SIDE, bottom=BORDER_SIDE,
    )

    def style_ws(ws, df):
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = CELL_BORDER
        for row_num in range(2, ws.max_row + 1):
            fill = ALT_FILL if row_num % 2 == 0 else PatternFill()
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill      = fill
                cell.border    = CELL_BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center")
        for col_num, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_num)
            max_len = max(
                len(str(col_name)),
                *[len(str(ws.cell(row=r, column=col_num).value or ""))
                  for r in range(2, ws.max_row + 1)],
                0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)
        ws.row_dimensions[1].height = 28

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # Sheet 1: Student Master
        df1 = pd.DataFrame([{
            "PRN":            s.prn,
            "Seat No":        s.seat_no,
            "Name":           s.name,
            "SGPA":           s.sgpa,
            "Credits Earned": s.credits_earned,
            "Credits Total":  s.credits_total,
            "Subjects":       len(s.subjects),
            "Status":         "PASS" if (s.sgpa or 0) >= 4.0 else "FAIL",
            "Category":       _sgpa_category(s.sgpa),
        } for s in students])
        df1.to_excel(writer, sheet_name="Student Master", index=False)
        style_ws(writer.sheets["Student Master"], df1)

        # Sheet 2: Rank List
        sorted_s = sorted(students, key=lambda s: s.sgpa or 0, reverse=True)
        df2 = pd.DataFrame([{
            "Rank":     i,
            "PRN":      s.prn,
            "Seat No":  s.seat_no,
            "Name":     s.name,
            "SGPA":     s.sgpa,
            "Status":   "PASS" if (s.sgpa or 0) >= 4.0 else "FAIL",
            "Category": _sgpa_category(s.sgpa),
        } for i, s in enumerate(sorted_s, 1)])
        df2.to_excel(writer, sheet_name="Rank List", index=False)
        style_ws(writer.sheets["Rank List"], df2)

        # Sheet 3: Subject Analytics
        subj_map = {}
        for s in students:
            for subj in s.subjects:
                subj_map.setdefault(subj.subject_code, []).append(subj)

        subj_rows = []
        for code, entries in sorted(subj_map.items()):
            totals   = [e.total for e in entries if e.total is not None]
            grades   = [e.grade for e in entries]
            passed   = sum(1 for g in grades if g not in ("F", "FF", "AB", None))
            appeared = len(entries)
            subj_rows.append({
                "Subject Code": code,
                "Appeared":     appeared,
                "Passed":       passed,
                "Failed":       appeared - passed,
                "Pass %":       round(passed / appeared * 100, 1) if appeared else 0,
                "Highest":      max(totals, default=0),
                "Lowest":       min(totals, default=0),
                "Average":      round(sum(totals) / len(totals), 2) if totals else 0,
            })
        df3 = pd.DataFrame(subj_rows)
        if not df3.empty:
            df3.to_excel(writer, sheet_name="Subject Analytics", index=False)
            style_ws(writer.sheets["Subject Analytics"], df3)

        # Sheet 4: SGPA Distribution
        bins   = [0, 4, 5, 6, 6.75, 7.75, 8.5, 10]
        labels = ["<4 (Fail)", "4-5 (Pass)", "5-6 (Second Class)",
                  "6-6.75 (Higher Second)", "6.75-7.75 (First Class)",
                  "7.75-8.5 (Distinction)", ">8.5 (Outstanding)"]
        sgpas = [s.sgpa for s in students if s.sgpa is not None]
        dist  = {lbl: 0 for lbl in labels}
        for sgpa in sgpas:
            for i in range(len(bins) - 1):
                if bins[i] <= sgpa < bins[i + 1]:
                    dist[labels[i]] += 1
                    break
            else:
                dist[labels[-1]] += 1
        df4 = pd.DataFrame(list(dist.items()), columns=["Category", "Count"])
        df4.to_excel(writer, sheet_name="SGPA Distribution", index=False)
        style_ws(writer.sheets["SGPA Distribution"], df4)

        # Sheet 5: Summary
        sgpa_vals  = [s.sgpa for s in students if s.sgpa is not None]
        pass_count = sum(1 for s in students if (s.sgpa or 0) >= 4.0)
        df5 = pd.DataFrame([
            ("University",              metadata.university_name),
            ("College",                 metadata.college_name),
            ("Department",              metadata.department_name),
            ("Semester",                metadata.semester_number),
            ("Session",                 f"{metadata.session_type} {metadata.session_year}"),
            ("", ""),
            ("Total Students",          len(students)),
            ("Average SGPA",            round(sum(sgpa_vals) / len(sgpa_vals), 2) if sgpa_vals else 0),
            ("Highest SGPA",            max(sgpa_vals, default=0)),
            ("Lowest SGPA",             min(sgpa_vals, default=0)),
            ("Pass Count",              pass_count),
            ("Fail Count",              len(students) - pass_count),
            ("Pass %",                  round(pass_count / len(students) * 100, 1) if students else 0),
            ("Distinctions (>=7.75)",   sum(1 for v in sgpa_vals if v >= 7.75)),
            ("First Class (6.75-7.75)", sum(1 for v in sgpa_vals if 6.75 <= v < 7.75)),
            ("Total Subjects",          len(subj_map)),
        ], columns=["Metric", "Value"])
        df5.to_excel(writer, sheet_name="Summary", index=False)
        style_ws(writer.sheets["Summary"], df5)

    output.seek(0)
    return output.read()


# ==============================================================================
# SIDEBAR  ← services are defined ABOVE this block, so they work everywhere
# ==============================================================================

with st.sidebar:

    # ── Logo ──────────────────────────────────────────────────────────────────
=======
# ── Theme ─────────────────────────────────────────────────────────────────────
from utils.theme import theme_manager

theme_manager.apply()

# ── Auth ──────────────────────────────────────────────────────────────────────
from utils.auth import auth_manager

# ── SVG icon helpers ──────────────────────────────────────────────────────────
_GITHUB_SVG = '<svg width="18" height="18" viewBox="0 0 24 24" fill="#ffffff"><path d="M12 0C5.37 0 0 5.37 0 12c0 5.31 3.435 9.795 8.205 11.385.6.105.825-.255.825-.57 0-.285-.015-1.23-.015-2.235-3.015.555-3.795-.735-4.035-1.41-.135-.345-.72-1.41-1.23-1.695-.42-.225-1.02-.78-.015-.795.945-.015 1.62.87 1.845 1.23 1.08 1.815 2.805 1.305 3.495.99.105-.78.42-1.305.765-1.605-2.67-.3-5.46-1.335-5.46-5.925 0-1.305.465-2.385 1.23-3.225-.12-.3-.54-1.53.12-3.18 0 0 1.005-.315 3.3 1.23.96-.27 1.98-.405 3-.405s2.04.135 3 .405c2.295-1.56 3.3-1.23 3.3-1.23.66 1.65.24 2.88.12 3.18.765.84 1.23 1.905 1.23 3.225 0 4.605-2.805 5.625-5.475 5.925.435.375.81 1.095.81 2.22 0 1.605-.015 2.895-.015 3.3 0 .315.225.69.825.57A12.02 12.02 0 0 0 24 12c0-6.63-5.37-12-12-12z"/></svg>'
_LINKEDIN_SVG = '<svg width="18" height="18" viewBox="0 0 24 24" fill="#ffffff"><path d="M20.447 20.452h-3.554v-5.569c0-1.328-.027-3.037-1.852-3.037-1.853 0-2.136 1.445-2.136 2.939v5.667H9.351V9h3.414v1.561h.046c.477-.9 1.637-1.85 3.37-1.85 3.601 0 4.267 2.37 4.267 5.455v6.286zM5.337 7.433a2.062 2.062 0 0 1-2.063-2.065 2.064 2.064 0 1 1 2.063 2.065zm1.782 13.019H3.555V9h3.564v11.452zM22.225 0H1.771C.792 0 0 .774 0 1.729v20.542C0 23.227.792 24 1.771 24h20.451C23.2 24 24 23.227 24 22.271V1.729C24 .774 23.2 0 22.222 0h.003z"/></svg>'
_INSTA_SVG = '<svg width="18" height="18" viewBox="0 0 24 24" fill="#ffffff"><path d="M12 2.163c3.204 0 3.584.012 4.85.07 3.252.148 4.771 1.691 4.919 4.919.058 1.265.069 1.645.069 4.849 0 3.205-.012 3.584-.069 4.849-.149 3.225-1.664 4.771-4.919 4.919-1.266.058-1.644.07-4.85.07-3.204 0-3.584-.012-4.849-.07-3.26-.149-4.771-1.699-4.919-4.92-.058-1.265-.07-1.644-.07-4.849 0-3.204.013-3.583.07-4.849.149-3.227 1.664-4.771 4.919-4.919 1.266-.057 1.645-.069 4.849-.069zM12 0C8.741 0 8.333.014 7.053.072 2.695.272.273 2.69.073 7.052.014 8.333 0 8.741 0 12c0 3.259.014 3.668.072 4.948.2 4.358 2.618 6.78 6.98 6.98C8.333 23.986 8.741 24 12 24c3.259 0 3.668-.014 4.948-.072 4.354-.2 6.782-2.618 6.979-6.98.059-1.28.073-1.689.073-4.948 0-3.259-.014-3.667-.072-4.947-.196-4.354-2.617-6.78-6.979-6.98C15.668.014 15.259 0 12 0zm0 5.838a6.162 6.162 0 1 0 0 12.324 6.162 6.162 0 0 0 0-12.324zM12 16a4 4 0 1 1 0-8 4 4 0 0 1 0 8zm6.406-11.845a1.44 1.44 0 1 0 0 2.881 1.44 1.44 0 0 0 0-2.881z"/></svg>'
_WEB_SVG = '<svg width="18" height="18" viewBox="0 0 24 24" fill="#ffffff"><path d="M12 0a12 12 0 1 0 0 24A12 12 0 0 0 12 0zm-1 17.93V16a1 1 0 0 0-1-1H6.07A10.01 10.01 0 0 1 2 12c0-.34.017-.68.05-1.01L6 15v1a2 2 0 0 0 2 2v.93zM17.9 17.39A2 2 0 0 0 16 16h-1v-3a1 1 0 0 0-1-1H8v-2h2a1 1 0 0 0 1-1V7h2a2 2 0 0 0 2-2v-.41A9.98 9.98 0 0 1 22 12a9.9 9.9 0 0 1-4.1 5.39z"/></svg>'


def _social_btn(url, bg, svg, label):
    return f"""
<a href="{url}" target="_blank" style="text-decoration:none;">
<div style="background:{bg}; border-radius:8px; padding:8px 12px; margin-bottom:7px;
            display:flex; align-items:center; gap:10px; cursor:pointer;">
  {svg}
  <span style="font-size:12px; color:#ffffff !important; font-weight:600;">{label}</span>
</div></a>"""


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:

    # Logo
>>>>>>> origin/develop
    logo_path = Path("logo.png")
    if logo_path.exists():
        st.image(str(logo_path), use_container_width=True)
        st.markdown(
            '<div style="text-align:center; font-size:11px; color:#7fa8d4; '
            'margin-top:-6px; margin-bottom:4px; letter-spacing:0.5px;">'
<<<<<<< HEAD
            'University Result Processing Platform</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown('<div class="sidebar-logo">🎓 ResultOps</div>', unsafe_allow_html=True)
=======
            "University Result Processing Platform</div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div class="sidebar-logo">🎓 ResultOps</div>', unsafe_allow_html=True
        )
>>>>>>> origin/develop
        st.markdown(
            '<div class="sidebar-tagline">University Result Processing Platform</div>',
            unsafe_allow_html=True,
        )

    st.markdown("---")

<<<<<<< HEAD
    # ── Navigation ────────────────────────────────────────────────────────────
    page = st.radio(
        "Navigation",
        ["📤  Upload Result", "📊  Dashboard", "📁  Export", "📋  History"],
=======
    # Navigation
    selected_page = st.radio(
        "Navigation",
        [
            "📤 Upload & Parse",
            "📊 Analytics Dashboard",
            "📋 History",
            "⚙️ System Stats",
        ],
>>>>>>> origin/develop
        label_visibility="collapsed",
    )

    st.markdown("---")

<<<<<<< HEAD
    # ── Credits / Social links ─────────────────────────────────────────────────
    st.markdown("---")

    # ── Version info ───────────────────────────────────────────────────────────
    st.markdown(
        '<div style="text-align:center; font-size:11px; color:#5a7fa8; '
        'margin-bottom:6px; letter-spacing:0.5px;">'
        'ResultOps v1.0 · Firebase Edition<br>'
        '</div>',
        unsafe_allow_html=True,
    )

    # ── Built by ───────────────────────────────────────────────────────────────
    st.markdown(
        '<div style="text-align:center; font-size:12px; color:#93b5e1; '
        'font-weight:600; margin-top:10px;">Built with ❤️ by</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div style="text-align:center; font-size:15px; color:#ffffff; '
        'font-weight:700; margin-bottom:12px;">Himanshu Jadhav</div>',
        unsafe_allow_html=True,
    )

    # ── GitHub ─────────────────────────────────────────────────────────────────
    st.markdown(
        '<a href="https://github.com/himanshu-jadhav108" target="_blank" '
        'style="text-decoration:none;">'
        '<div style="background:#1a1a1a; border:1px solid #444; border-radius:7px; '
        'padding:7px 12px; margin-bottom:7px; display:flex; align-items:center; gap:8px;">'
        '<span style="font-size:11px; color:#e8f0fe; font-weight:500;">GitHub Profile</span>'
        '</div></a>',
        unsafe_allow_html=True,
    )

    # ── LinkedIn ───────────────────────────────────────────────────────────────
    st.markdown(
        '<a href="https://www.linkedin.com/in/himanshu-jadhav-328082339" target="_blank" '
        'style="text-decoration:none;">'
        '<div style="background:#0a66c2; border-radius:7px; '
        'padding:7px 12px; margin-bottom:7px; display:flex; align-items:center; gap:8px;">'
        '<span style="font-size:11px; color:#ffffff; font-weight:500;">LinkedIn</span>'
        '</div></a>',
        unsafe_allow_html=True,
    )

    # ── Instagram ──────────────────────────────────────────────────────────────
    st.markdown(
        '<a href="https://www.instagram.com/himanshu_jadhav_108" target="_blank" '
        'style="text-decoration:none;">'
        '<div style="background:linear-gradient(135deg,#833ab4,#fd1d1d,#fcb045); '
        'border-radius:7px; padding:7px 12px; margin-bottom:7px; '
        'display:flex; align-items:center; gap:8px;">'
        '<span style="font-size:11px; color:#ffffff; font-weight:500;">Instagram</span>'
        '</div></a>',
        unsafe_allow_html=True,
    )

    # ── Portfolio ──────────────────────────────────────────────────────────────
    st.markdown(
        '<a href="https://himanshu-jadhav-portfolio.vercel.app/" target="_blank" '
        'style="text-decoration:none;">'
        '<div style="background:linear-gradient(135deg,#1a4a8a,#2d8cff); '
        'border-radius:7px; padding:7px 12px; margin-bottom:12px; '
        'display:flex; align-items:center; gap:8px;">'
        '<span style="font-size:16px;">🌐</span>'
        '<span style="font-size:11px; color:#ffffff; font-weight:500;">Portfolio</span>'
        '</div></a>',
        unsafe_allow_html=True,
    )

    # ── Copyright ──────────────────────────────────────────────────────────────
    st.markdown(
        '<div style="text-align:center; font-size:10px; color:#3d5a7a; '
        'margin-top:4px;">© 2025 ResultOps</div>',
        unsafe_allow_html=True,
    )

# ==============================================================================
# PAGE: UPLOAD RESULT
# ==============================================================================

def page_upload():
    st.title("📤 Upload University Result PDF")
    st.markdown(
        "Upload an **official university ledger PDF** (text-based, not scanned). "
        "ResultOps auto-detects all metadata and parses every student record."
    )
    st.markdown("---")

    uploaded_file = st.file_uploader(
        "📂 Choose PDF file",
        type=["pdf"],
        help="Must be a text-based university ledger PDF (not a scanned image).",
    )

    if not uploaded_file:
        st.markdown("""
        <div style="text-align:center; padding:48px; color:#5a7fa8;">
            <div style="font-size:52px; margin-bottom:12px;">📄</div>
            <div style="font-size:18px; font-weight:600; color:#93b5e1;">No file uploaded yet</div>
            <div style="font-size:13px; margin-top:8px;">
                Drag & drop or click Browse above to get started
            </div>
        </div>
        """, unsafe_allow_html=True)
        return

    pdf_bytes = uploaded_file.read()

    # ── Parse ─────────────────────────────────────────────────────────────────
    progress = st.progress(0, text="Starting...")
    try:
        from parser.pdf_parser import extract_text_from_pdf
        from parser.metadata_extractor import extract_metadata
        from parser.student_parser import parse_students
        from utils.validators import validate_students

        progress.progress(20, text="📖 Extracting text from PDF...")
        full_text = extract_text_from_pdf(pdf_bytes)

        progress.progress(50, text="🔍 Detecting metadata...")
        metadata = extract_metadata(full_text)

        progress.progress(75, text="👥 Parsing student records...")
        students = parse_students(full_text)

        progress.progress(100, text="✅ Done!")
        time.sleep(0.3)
        progress.empty()

    except Exception as e:
        progress.empty()
        st.error(f"❌ Parsing failed: {e}")
        logger.exception("Parsing error")
        return

    # ── Metadata cards — full text, no truncation ──────────────────────────────
    st.markdown("### 🔍 Detected Metadata")
    c1, c2, c3, c4, c5 = st.columns(5)

    with c1:
        st.markdown(f"""
        <div class="meta-card" style="border-top:3px solid #2d8cff;">
            <div class="meta-label">🏛️ University</div>
            <div class="meta-value">{metadata.university_name}</div>
        </div>""", unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
        <div class="meta-card" style="border-top:3px solid #2d8cff;">
            <div class="meta-label">🏫 College</div>
            <div class="meta-value">{metadata.college_name}</div>
        </div>""", unsafe_allow_html=True)

    with c3:
        st.markdown(f"""
        <div class="meta-card" style="border-top:3px solid #f5a623;">
            <div class="meta-label">📚 Department</div>
            <div class="meta-value">{metadata.department_name}</div>
        </div>""", unsafe_allow_html=True)

    with c4:
        st.markdown(f"""
        <div class="meta-card" style="border-top:3px solid #00d97e;">
            <div class="meta-label">📋 Semester</div>
            <div class="meta-value">Semester {metadata.semester_number}</div>
        </div>""", unsafe_allow_html=True)

    with c5:
        st.markdown(f"""
        <div class="meta-card" style="border-top:3px solid #a78bfa;">
            <div class="meta-label">📅 Session</div>
            <div class="meta-value">{metadata.session_type} {metadata.session_year}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)
    st.markdown("---")

    # ── Validation ─────────────────────────────────────────────────────────────
    st.markdown("### 🧪 Validation Report")
    validation = validate_students(students, expected_semester=metadata.semester_number)

    v1, v2 = st.columns([3, 2])
    with v1:
        for line in validation.summary_lines():
            if line.startswith("✅"):
                st.success(line)
            elif line.startswith("⚠️"):
                st.warning(line)
            elif line.startswith("❌"):
                st.error(line)

    with v2:
        if validation.is_valid:
            st.markdown("""
            <div style="background:#0a2e1a; border:1px solid #00d97e; border-radius:12px;
                        padding:28px; text-align:center;">
                <div style="font-size:40px;">✅</div>
                <div style="font-size:17px; font-weight:700; color:#00d97e; margin-top:10px;">
                    Validation Passed
                </div>
                <div style="color:#4dbb88; margin-top:6px; font-size:13px;">
                    Ready to save to database
                </div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown("""
            <div style="background:#2e0a0a; border:1px solid #ff4d6d; border-radius:12px;
                        padding:28px; text-align:center;">
                <div style="font-size:40px;">❌</div>
                <div style="font-size:17px; font-weight:700; color:#ff4d6d; margin-top:10px;">
                    Validation Failed
                </div>
                <div style="color:#cc6677; margin-top:6px; font-size:13px;">
                    Fix errors before saving
                </div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

    # ── Student Preview ────────────────────────────────────────────────────────
    if students:
        st.markdown("---")
        st.markdown(f"### 👥 Student Preview — {len(students)} records detected")
        preview_rows = [{
            "PRN":      s.prn,
            "Seat No":  s.seat_no,
            "Name":     s.name,
            "SGPA":     s.sgpa,
            "Subjects": len(s.subjects),
            "Status":   "✅ PASS" if (s.sgpa or 0) >= 4.0 else "❌ FAIL",
        } for s in students[:15]]
        st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)
        if len(students) > 15:
            st.caption(f"Showing first 15 of {len(students)} students.")

    # ── Download before save ───────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Download Analysed Report")
    st.caption("Get the full Excel report now — you don't need to save to database first.")

    dl_left, dl_right = st.columns([2, 3])
    with dl_left:
        try:
            excel_data = _generate_excel_report(metadata, students)
            st.download_button(
                label="⬇️ Download Excel Report (5 Sheets)",
                data=excel_data,
                file_name=_excel_filename(metadata),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.warning(f"Could not generate report: {e}")

    with dl_right:
        st.markdown("""
        <div style="background:#0d1f3c; border:1px solid rgba(45,140,255,0.2);
                    border-radius:10px; padding:14px 18px; font-size:13px; color:#93b5e1;">
            <b style="color:#e8f0fe;">Report includes 5 sheets:</b><br>
            📋 Student Master &nbsp;·&nbsp; 🏆 Rank List &nbsp;·&nbsp;
            📚 Subject Analytics<br>📊 SGPA Distribution &nbsp;·&nbsp; 📝 Summary
        </div>""", unsafe_allow_html=True)

    # ── Save to DB ─────────────────────────────────────────────────────────────
    st.markdown("---")

    if not validation.is_valid:
        st.error("⛔ Cannot save — fix the validation errors above first.")
        return

    save_col, note_col = st.columns([2, 3])
    with save_col:
        save_clicked = st.button(
            "💾 Save to ResultOps Database",
            type="primary",
            use_container_width=True,
        )
    with note_col:
        st.caption(
            "Saves all student records and marks to Firebase Firestore. "
            "Duplicate semester uploads are automatically blocked."
        )

    if save_clicked:
        from services.result_service import ResultService, DuplicateSemesterError
        try:
            with st.spinner("💾 Saving to Firebase — please wait..."):
                service = ResultService()
                summary = service.save_results(metadata, students)

            st.balloons()
            st.success("✅ Results saved successfully to Firebase!")

            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Students Saved", summary["students_inserted"])
            r2.metric("Results Saved",  summary["results_inserted"])
            r3.metric("Marks Saved",    summary["marks_inserted"])
            r4.metric("Semester",       f"Sem {summary['semester']}")

            st.markdown("---")
            st.markdown("### 📥 Download Saved Report")
            st.download_button(
                label="⬇️ Download Full Excel Report",
                data=_generate_excel_report(metadata, students),
                file_name=_excel_filename(metadata),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

        except DuplicateSemesterError as e:
            st.error(f"🚫 Duplicate Upload Blocked — {e}")
            st.markdown("---")
            st.markdown("### 📥 Download Report Anyway")
            st.download_button(
                label="⬇️ Download Excel Report",
                data=_generate_excel_report(metadata, students),
                file_name=_excel_filename(metadata),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"❌ Database error: {e}")
            logger.exception("DB save error")


# ==============================================================================
# PAGE: DASHBOARD
# ==============================================================================

def page_dashboard():
    st.title("📊 Analytics Dashboard")

    analytics, _ = try_load_services()
    if analytics is None:
        return

    st.markdown("### 🔎 Select Data")
    f1, f2, f3, f4 = st.columns(4)

    universities = analytics.get_universities()
    uni_list     = [u["name"] for u in universities]
    uni_select   = f1.selectbox("🏛️ University", ["— All —"] + uni_list)
    uni_filter   = uni_select if uni_select != "— All —" else None

    colleges   = analytics.get_colleges(uni_filter)
    col_list   = [c["name"] for c in colleges]
    col_select = f2.selectbox("🏫 College", ["— All —"] + col_list)
    col_filter = col_select if col_select != "— All —" else None

    departments = analytics.get_departments(col_filter)
    dept_list   = [d["name"] for d in departments]

    if not dept_list:
        st.info("📂 No departments found. Upload a result PDF first.")
        return

    dept_select = f3.selectbox("📚 Department", dept_list)

    semester_key = None
    rank_df      = pd.DataFrame()   # initialise so it's always defined
    semesters    = analytics.get_semesters_for_department(dept_select)
    if semesters:
        sem_labels = {
            f"Semester {s['semester_number']} — {s.get('session_type','')} {s.get('session_year','')}": s["id"]
            for s in semesters
        }
        sem_select   = f4.selectbox("📋 Semester", list(sem_labels.keys()))
        semester_key = sem_labels.get(sem_select)

    if not semester_key:
        st.info("📋 Select a semester to view analytics.")
        return

    st.markdown("---")

    summary = analytics.semester_summary(semester_key)
    if not summary:
        st.warning("No data found for this semester.")
        return

    st.markdown("### 📈 Semester Overview")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("👥 Students",     summary["total_students"])
    m2.metric("📊 Avg SGPA",     summary["avg_sgpa"])
    m3.metric("⭐ Highest SGPA", summary["max_sgpa"])
    m4.metric("🏆 Distinctions", summary["distinctions"])
    m5.metric("✅ Passed",        summary["pass_count"])
    m6.metric("📈 Pass %",        f"{summary['pass_percentage']}%")

    st.markdown("---")

    chart_col, top_col = st.columns([3, 2])

    with chart_col:
        st.markdown("### 📉 SGPA Distribution")
        dist_df = analytics.sgpa_distribution(semester_key)
        if not dist_df.empty:
            st.bar_chart(
                dist_df.set_index("SGPA Range")["Count"],
                color="#2d8cff",
                use_container_width=True,
            )

    with top_col:
        st.markdown("### 🏆 Top 10 Students")
        rank_df = analytics.student_rank_list(semester_key)
        if not rank_df.empty:
            st.dataframe(
                rank_df.head(10)[["Rank", "Name", "SGPA", "Category"]],
                use_container_width=True,
                hide_index=True,
            )

    st.markdown("---")

    st.markdown("### 📚 Subject-wise Analytics")
    subj_df = analytics.subject_analytics(semester_key)
    if not subj_df.empty:
        def color_pass(val):
            try:
                v = float(val)
                if v >= 80: return "background-color:#0a2e1a; color:#00d97e; font-weight:600"
                if v >= 60: return "background-color:#2e2000; color:#f5a623; font-weight:600"
                return "background-color:#3d0012; color:#ff4d6d; font-weight:600"
            except:
                return ""
        st.dataframe(
            subj_df.style.map(color_pass, subset=["Pass %"]),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.info("No subject data available.")

    # ── Full Rank List + Fail List ─────────────────────────────────────────────
    if not rank_df.empty:
        st.markdown("---")
        with st.expander("📋 Full Rank List — click to expand"):
            def highlight_row(row):
                if row["Status"] == "FAIL":
                    return ["background-color:#3d0012; color:#ff4d6d"] * len(row)
                if row["Category"] == "Distinction":
                    return ["background-color:#0a2e1a; color:#00d97e"] * len(row)
                if row["Category"] == "First Class":
                    return ["background-color:#0a1f3a; color:#2d8cff"] * len(row)
                return [""] * len(row)
            st.dataframe(
                rank_df.style.apply(highlight_row, axis=1),
                use_container_width=True,
                hide_index=True,
            )

        fail_df = rank_df[rank_df["Status"] == "FAIL"]
        if not fail_df.empty:
            with st.expander(f"⚠️ Fail List — {len(fail_df)} students"):
                st.dataframe(
                    fail_df[["Rank", "PRN", "Seat No", "Name", "SGPA"]],
                    use_container_width=True,
                    hide_index=True,
                )


# ==============================================================================
# PAGE: EXPORT
# ==============================================================================

def page_export():
    st.title("📁 Export Results")
    st.markdown("Generate a styled Excel workbook from any uploaded semester.")
    st.markdown("---")

    analytics, export_svc = try_load_services()
    if analytics is None:
        return

    departments = analytics.get_departments()
    if not departments:
        st.info("No data available. Upload a result PDF first.")
        return

    e1, e2 = st.columns(2)
    dept_list   = [d["name"] for d in departments]
    dept_select = e1.selectbox("📚 Department", dept_list)

    semester_key = None
    sem_label    = ""
    semesters    = analytics.get_semesters_for_department(dept_select)
    if semesters:
        sem_labels = {
            f"Semester {s['semester_number']} — {s.get('session_type','')} {s.get('session_year','')}": s["id"]
            for s in semesters
        }
        sem_select   = e2.selectbox("📋 Semester", list(sem_labels.keys()))
        semester_key = sem_labels.get(sem_select)
        sem_label    = sem_select

    if not semester_key:
        return

    st.markdown("---")
    st.markdown("### 📊 Quick Preview")
    summary = analytics.semester_summary(semester_key)
    if summary:
        p1, p2, p3, p4 = st.columns(4)
        p1.metric("👥 Students",     summary["total_students"])
        p2.metric("📊 Avg SGPA",     summary["avg_sgpa"])
        p3.metric("📈 Pass %",        f"{summary['pass_percentage']}%")
        p4.metric("🏆 Distinctions", summary["distinctions"])

    st.markdown("---")
    btn_col, _ = st.columns([2, 3])
    with btn_col:
        if st.button("📥 Generate Excel Report", type="primary", use_container_width=True):
            with st.spinner("Generating Excel workbook..."):
                try:
                    excel_bytes = export_svc.generate_excel(semester_key)
                    safe_dept   = dept_select.replace(" ", "_")[:20]
                    safe_sem    = sem_label.replace(" ", "_").replace("—", "").strip()[:15]
                    filename    = f"ResultOps_{safe_dept}_{safe_sem}.xlsx"
                    st.download_button(
                        label="⬇️ Download Excel Report",
                        data=excel_bytes,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                    st.success("✅ Ready! Click the button above to download.")
                except Exception as e:
                    st.error(f"Export failed: {e}")


# ==============================================================================
# PAGE: HISTORY
# ==============================================================================

def page_history():
    st.title("📋 Upload History")
    st.markdown("All previously uploaded semester results.")
    st.markdown("---")

    analytics, _ = try_load_services()
    if analytics is None:
        return

    with st.spinner("Loading history..."):
        history_df = analytics.list_uploaded_semesters()

    if history_df.empty:
        st.markdown("""
        <div style="text-align:center; padding:60px; color:#5a7fa8;">
            <div style="font-size:52px; margin-bottom:12px;">📭</div>
            <div style="font-size:18px; font-weight:600; color:#93b5e1;">No uploads yet</div>
            <div style="font-size:13px; margin-top:8px;">
                Go to Upload Result to add your first semester
            </div>
        </div>""", unsafe_allow_html=True)
        return

    h1, h2, h3 = st.columns(3)
    h1.metric("📋 Semesters Uploaded", len(history_df))
    h2.metric("📚 Departments",         history_df["Department"].nunique())
    h3.metric("🏛️ Universities",        history_df["University"].nunique())

    st.markdown("---")
    st.dataframe(
        history_df.drop(columns=["Semester Key"]),
        use_container_width=True,
        hide_index=True,
    )

    # Admin delete
    st.markdown("---")
    with st.expander("🔴 Admin: Delete a Semester (Irreversible)"):
        st.warning("⚠️ This permanently deletes all student records. Cannot be undone.")
        admin_pw    = st.text_input("🔐 Admin Password", type="password", key="admin_pw")
        expected_pw = os.environ.get("ADMIN_PASSWORD", "")

        if admin_pw and expected_pw and admin_pw == expected_pw:
            st.success("✅ Access granted")

            def fmt_sem(sid):
                row = history_df[history_df["Semester Key"] == sid]
                if row.empty: return sid
                r = row.iloc[0]
                return f"{r['Department']} | Sem {r['Semester No']} | {r['Session']}"

            semester_to_delete = st.selectbox(
                "Select Semester to Delete",
                history_df["Semester Key"].tolist(),
                format_func=fmt_sem,
            )
            del_col, _ = st.columns([2, 3])
            with del_col:
                if st.button("🗑️ Confirm Delete", type="secondary", use_container_width=True):
                    try:
                        analytics.delete_semester(semester_to_delete)
                        st.success("Semester deleted successfully.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Delete failed: {e}")

        elif admin_pw:
            st.error("❌ Incorrect admin password.")


# ==============================================================================
# ROUTER
# ==============================================================================

if page == "📤  Upload Result":
    page_upload()
elif page == "📊  Dashboard":
    page_dashboard()
elif page == "📁  Export":
    page_export()
elif page == "📋  History":
    page_history()
=======
    # Theme toggle
    theme_manager.render_toggle()

    st.markdown("---")

    # Auth status
    if auth_manager.is_admin_authenticated:
        st.success("🔒 ADMIN Access")
    if auth_manager.is_write_authenticated:
        st.success("✅ WRITE Access")
    if auth_manager.is_read_authenticated:
        st.success("✅ READ Access")

    auth_manager.render_logout_button()

    st.markdown("---")

    # Footer
    st.markdown(
        '<div style="text-align:center; font-size:11px; color:#5a7fa8; '
        'margin-bottom:4px; letter-spacing:0.5px;">'
        "ResultOps v2.0 · Firebase Edition</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div style="text-align:center; font-size:12px; color:#93b5e1; font-weight:600; margin-top:8px;">'
        "Built with ❤️ by</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div style="text-align:center; font-size:15px; color:#ffffff; font-weight:700; margin-bottom:10px;">'
        "Himanshu Jadhav</div>",
        unsafe_allow_html=True,
    )

    st.markdown(
        _social_btn(
            "https://github.com/himanshu-jadhav108",
            "#1a1a1a; border:1px solid #555",
            _GITHUB_SVG,
            "GitHub Profile",
        )
        + _social_btn(
            "https://www.linkedin.com/in/himanshu-jadhav-328082339",
            "#0a66c2",
            _LINKEDIN_SVG,
            "LinkedIn",
        )
        + _social_btn(
            "https://www.instagram.com/himanshu_jadhav_108",
            "linear-gradient(135deg,#833ab4 0%,#fd1d1d 50%,#fcb045 100%)",
            _INSTA_SVG,
            "Instagram",
        )
        + _social_btn(
            "https://himanshu-jadhav-portfolio.vercel.app/",
            "linear-gradient(135deg,#1a4a8a,#2d8cff)",
            _WEB_SVG,
            "Portfolio",
        ),
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div style="text-align:center; font-size:10px; color:#3d5a7a; margin-top:6px;">'
        "© 2025 ResultOps</div>",
        unsafe_allow_html=True,
    )

# ── Route to page modules ─────────────────────────────────────────────────────
if selected_page == "📤 Upload & Parse":
    from views import upload_page

    upload_page.render()

elif selected_page == "📊 Analytics Dashboard":
    if auth_manager.require_read_auth():
        from views import analytics_page

        analytics_page.render()

elif selected_page == "📋 History":
    if auth_manager.require_read_auth():
        from views import history_page

        history_page.render()

elif selected_page == "⚙️ System Stats":
    if auth_manager.require_admin_auth():
        from views import system_stats

        system_stats.render()
>>>>>>> origin/develop
